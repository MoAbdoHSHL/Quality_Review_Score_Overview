## 
## By Mohamed Abdo </>
## 

import subprocess as sp
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from io import BytesIO
from urllib.parse import quote
import os
import re
import subprocess
import concurrent.futures
import pandas as pd

def encode_url(url):
    return quote(url, safe=":/")

def fetch_svn_info(svn_url):
    try:
        encoded_url = encode_url(svn_url)
        svn_info_command = ['svn', 'info', encoded_url]
        result = subprocess.run(svn_info_command, capture_output=True, text=True)
        
        if result.returncode != 0:
            print(f"Error: Unable to fetch SVN info for '{svn_url}'.\n{result.stderr}")
            return None
        
        info = {}
        for line in result.stdout.splitlines():
            if line.startswith("Revision:"):
                info["revision"] = line.split(":")[1].strip()
            elif line.startswith("Last Changed Date:"):
                date_info = line.split(":", 1)[1].strip()
                parts = date_info.split()
                if len(parts) >= 2:
                    info["last_changed_date"] = f"{parts[0]} {parts[1]}"
                else:
                    info["last_changed_date"] = date_info
            elif line.startswith("Last Changed Rev:"):
                info["last_changed_rev"] = line.split(":")[1].strip()
        return info
    except Exception as e:
        print(f"An error occurred while fetching the SVN info: {e}")
        return None

def fetch_file_from_svn(svn_file_url):
    try:
        encoded_url = encode_url(svn_file_url)
        svn_cat_command = ['svn', 'cat', encoded_url]
        result = subprocess.run(svn_cat_command, capture_output=True)
        
        if result.returncode != 0:
            print(f"Error: Unable to fetch file '{svn_file_url}'.\n{result.stderr}")
            return None
        
        return BytesIO(result.stdout)
    except Exception as e:
        print(f"An error occurred while fetching the file from SVN: {e}")
        return None

def process_xlsx_file(file_stream):
    try:
        wb_input = openpyxl.load_workbook(file_stream, data_only=True)
        if "DR-SW" in wb_input.sheetnames:
            sheet_input = wb_input["DR-SW"]
            cell_value = sheet_input["K7"].value  # Actual Version Reviewed
            review_score = sheet_input["J2"].value  # Review Score
            
            # Extract numeric part from review_score using regex
            match = re.match(r"([\d.]+)", str(review_score).strip())
            if match:
                review_score = float(match.group(1))
            else:
                review_score = None
            
            actual_version_reviewed = cell_value
            if actual_version_reviewed is not None:
                actual_version_reviewed = str(actual_version_reviewed).strip()
            
            # Return extracted values
            return {
                "actual_version_reviewed": actual_version_reviewed,
                "review_score": review_score
            }

    except Exception as e:
        print(f"An error occurred while processing the file: {e}")
        return None

def generate_output_excel(data, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Review Board"
    
    headers = [
        "SVN Link Review File", "Actual Version Reviewed", "Last Changed Date",
        "File Name", "Folder Name", "Main Controller", "Last Changed Rev", "Review Score"
    ]
    
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    header_font = Font(bold=True, size=14)
    
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    rows = []

    for row in data:
        review_score = row.get('review_score', 0)
        formatted_review_score = f"{review_score * 100:.2f}%" if review_score is not None else "N/A"
        last_changed_date = row.get('last_changed_date', '')
        try:
            if last_changed_date:
                # Convert string to datetime and format
                date_obj = pd.to_datetime(last_changed_date, errors='coerce')
                last_changed_date = date_obj.strftime('%Y-%m-%d') if date_obj else ''
        except Exception as e:
            print(f"Error formatting date '{last_changed_date}': {e}")
            last_changed_date = ''
        
        ws.append([
            row["file_path"], 
            row.get("actual_version_reviewed", "N/A") or "N/A",  
            last_changed_date,
            row["file_name"], 
            row["folder_name"], 
            row["main_controller"],
            row.get("last_changed_rev", ""),  
            formatted_review_score
        ])

        rows.append([
            row["file_path"], 
            row.get("actual_version_reviewed", "N/A") or "N/A", 
            last_changed_date,
            row["file_name"], 
            row["folder_name"], 
            row["main_controller"],
            row.get("last_changed_rev", ""),  
            formatted_review_score
        ])
        
    column_widths = {
        'A': 30, 
        'B': 28, 
        'C': 25, 
        'D': 62, 
        'E': 25, 
        'F': 23, 
        'G': 23, 
        'H': 20
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))

    # Center-align all cells except the first column (left-align)
    for row in ws.iter_rows(min_row=2):  # Skip header row
        for cell in row:
            if cell.column_letter == 'A':
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_style

    # Apply Color formatting
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    
    review_score_font = Font(size=14, bold=True)
    
    ws.auto_filter.ref = ws.dimensions  #filter
    
    # Re-sort the rows by "Last Changed Date" column from newest to oldest
    ws.auto_filter.ref = ws.dimensions  # Apply filter to the whole table
    sorted_rows = sorted(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=8), key=lambda r: r[2].value, reverse=True)
    
    for idx, row in enumerate(sorted_rows, start=2):
        for col, cell in enumerate(row, start=1):
            ws.cell(row=idx, column=col, value=cell.value)

    # Apply conditional formatting to the Review Score column
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                try:
                    # Extract numeric score
                    score_value = float(cell.value.strip('%'))
                    if score_value < 80:
                        cell.fill = red_fill
                    elif 80 <= score_value < 95:
                        cell.fill = yellow_fill
                    elif score_value >= 95:
                        cell.fill = green_fill
                    cell.font = review_score_font  # Set font size and bold
                except (ValueError, TypeError):
                    # In case cell.value is None or invalid, just skip
                    continue

    overview_ws = wb.create_sheet(title="Overview")
    
    # Define headers 
    overview_headers = ["Score Range", "Count", "Percentage"]
    overview_ws.append(overview_headers)
    
    header_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Define score ranges
    ranges = [
        ("<80%", 0, 80),
        ("80% - 95%", 80, 95),
        (">=95%", 95, 100)
    ]
    
    total_count = len(data)
    count_dict = {r[0]: 0 for r in ranges}
    
    # Calculate counts for each range
    for row in data:
        try:
            score_value = float(row.get('review_score', 0) * 100)
            if score_value < 80:
                count_dict["<80%"] += 1
            elif 80 <= score_value < 95:
                count_dict["80% - 95%"] += 1
            elif score_value >= 95:
                count_dict[">=95%"] += 1
        except (ValueError, TypeError):
            # Skip invalid scores
            continue

    # Populate overview data
    for label, min_val, max_val in ranges:
        count = count_dict[label]
        percentage = (count / total_count * 100) if total_count > 0 else 0
        overview_ws.append([label, count, f"{percentage:.2f}%"])
    
    for cell in overview_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in overview_ws.iter_rows(min_row=2, max_row=overview_ws.max_row, max_col=3):
        for cell in row:
            cell.border = border_style
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=14, bold=True)
    
    for row in overview_ws.iter_rows(min_row=2, max_row=4, min_col=1, max_col=1):
        for cell in row:
            score_label = cell.value
            if score_label == "<80%":
                cell.fill = red_fill
            elif score_label == "80% - 95%":
                cell.fill = yellow_fill
            elif score_label == ">=95%":
                cell.fill = green_fill
        cell.font = Font(size=14, bold=True)

    # Create a pie chart 
    pie_chart = PieChart()
    labels = Reference(overview_ws, min_col=1, min_row=2, max_row=4)
    data = Reference(overview_ws, min_col=2, min_row=1, max_row=4)  
    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(labels)
    pie_chart.dataLabels = DataLabelList()
    pie_chart.dataLabels.showPercent = True
    pie_chart.dataLabels.showLabel = False
    pie_chart.dataLabels.showLeaderLines = False

    pie_chart.dataLabels.font = Font(size=14, bold=True)

    overview_ws.add_chart(pie_chart, "A7")  
    overview_ws.column_dimensions['A'].width = 30
    overview_ws.column_dimensions['B'].width = 20
    overview_ws.column_dimensions['C'].width = 30

    
    wb.save(output_path)

    print(f"Data successfully saved to '{output_path}'.")
    try:
        sp.Popen(["start", output_path], shell=True)
    except Exception as e:
        print(f"An error occurred while opening the file: {e}")



def recursively_list_files(svn_url):
    all_files = []
    directories_to_check = [svn_url]
    
    while directories_to_check:
        current_url = directories_to_check.pop()
        listed_files = list_svn_files(current_url)
        for file_name in listed_files:
            if file_name.endswith('/'):
                # Directory - add it to the list to be checked
                directories_to_check.append(f"{current_url}/{file_name}")
            elif file_name.endswith('.xlsx'):
                # File - add it to the list of files
                all_files.append(f"{current_url}/{file_name}")
    
    return all_files

def list_svn_files(svn_url):
    try:
        encoded_url = encode_url(svn_url)
        svn_list_command = ['svn', 'list', encoded_url]
        result = subprocess.run(svn_list_command, capture_output=True, text=True)
        
        if result.returncode != 0:
            print(f"Error: Unable to list files for '{svn_url}'.\n{result.stderr}")
            return []
        
        files = result.stdout.splitlines()
        return files
    except Exception as e:
        print(f"An error occurred while listing files: {e}")
        return []

def process_files(file_paths):
    """Process multiple files concurrently."""
    all_data = []
    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = {executor.submit(fetch_file_from_svn, path): path for path in file_paths}
        for future in concurrent.futures.as_completed(futures):
            file_path = futures[future]
            file_stream = future.result()
            if file_stream:
                result = process_xlsx_file(file_stream)
                if result:
                    # Fetch SVN info for each individual file
                    file_info = fetch_svn_info(file_path)
                    all_data.append({
                        "file_path": file_path,
                        "file_name": os.path.basename(file_path),
                        "folder_name": os.path.dirname(file_path).split('/')[-1],
                        "actual_version_reviewed": result["actual_version_reviewed"],
                        "review_score": result["review_score"],
                        "main_controller": "ComController DSPA" if "01_ComController DSPA" in file_path else "ObcController DSPB",
                        "last_changed_rev": file_info.get("last_changed_rev", "") if file_info else "",
                        "last_changed_date": file_info.get("last_changed_date", "") if file_info else ""
                    })
    return all_data


def main():
    svn_urls = [
        "https:link 1",
        "https:link 2"
    ]
    
    output_file_path = "D:output.xlsx"
    
    all_data = []

    for svn_url in svn_urls:
        # No longer fetching info at the folder level
        file_list = recursively_list_files(svn_url)
        file_data = process_files(file_list)
        
        all_data.extend(file_data)
    
    generate_output_excel(all_data, output_file_path)

if __name__ == "__main__":
    main()
